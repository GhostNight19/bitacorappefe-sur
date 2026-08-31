# -*- coding: utf-8 -*-
"""
Conversor del GRAFICO DE TURNOS (Excel) -> grafico/grafico.json

Uso:
    python convertir_grafico.py

Deja el Excel del grafico del mes dentro de la carpeta "graficos_excel" y
ejecuta el script. Genera "grafico/grafico.json", que es lo que lee la
aplicacion (y lo que hay que subir al repositorio).

El Excel trae tres hojas y las tres se conservan tal cual:
  - "Grafico <mes>"   : quien hace que turno cada dia
  - "Lunes - Viernes" : catalogo de horarios de los turnos de dia de semana
  - "S-D-F"           : catalogo de sabado, domingo y feriado

Igual que con las pautas y el boletin, el script no interpreta nada: vuelca las
filas como texto y la lectura la hace el propio index.html.
"""

import json
import re
import unicodedata
from datetime import date, datetime, time, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    raise SystemExit("Falta openpyxl. Instalalo con:  pip install openpyxl")

BASE = Path(__file__).resolve().parent
ORIGEN = BASE / "graficos_excel"
DESTINO = BASE / "grafico" / "grafico.json"

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}


def sin_acentos(texto):
    return "".join(
        c for c in unicodedata.normalize("NFD", str(texto))
        if unicodedata.category(c) != "Mn"
    )


def celda_a_texto(valor):
    """Texto plano. Las horas quedan HH:MM y las fechas AAAA-MM-DD."""
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "SI" if valor else "NO"
    if isinstance(valor, datetime):
        if valor.year > 1900:
            return valor.strftime("%Y-%m-%d")
        return valor.strftime("%H:%M")
    if isinstance(valor, time):
        return valor.strftime("%H:%M")
    if isinstance(valor, timedelta):
        total = int(valor.total_seconds() // 60)
        return "%d:%02d" % (total // 60, total % 60)
    if isinstance(valor, date):
        return valor.isoformat()
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).replace("\n", " ").strip()


def leer_hoja(ws):
    filas = []
    for fila in ws.iter_rows(values_only=True):
        celdas = [celda_a_texto(v) for v in fila]
        while celdas and celdas[-1] == "":
            celdas.pop()
        filas.append(celdas)
    while filas and not any(filas[-1]):
        filas.pop()
    return filas


def mes_del_grafico(hojas, nombre_archivo):
    """'Grafico Septiembre' + la primera fecha -> 2026-09."""
    for filas in hojas.values():
        for fila in filas[:40]:
            for celda in fila:
                m = re.match(r"(20\d{2})-(\d{2})-\d{2}", str(celda))
                if m:
                    return "%s-%s" % (m.group(1), m.group(2))
    texto = sin_acentos(nombre_archivo).lower()
    for nombre, num in MESES.items():
        if nombre in texto:
            anio = re.search(r"(20\d{2})", texto)
            return "%s-%02d" % (anio.group(1) if anio else date.today().year, num)
    return ""


def main():
    if not ORIGEN.exists():
        ORIGEN.mkdir(parents=True)
        print('Se creo la carpeta "graficos_excel". Deja ahi el grafico y vuelve a ejecutar.')
        return

    archivos = sorted(
        p for p in ORIGEN.iterdir()
        if p.suffix.lower() in (".xlsx", ".xlsm") and not p.name.startswith("~$")
    )
    if not archivos:
        print('No hay Excel en "graficos_excel". Deja ahi el grafico y vuelve a ejecutar.')
        return

    graficos = []
    for ruta in archivos:
        wb = openpyxl.load_workbook(ruta, data_only=True)
        hojas = {}
        principal = ""
        for nombre in wb.sheetnames:
            filas = leer_hoja(wb[nombre])
            clave = sin_acentos(nombre).strip().upper()
            if clave.startswith("GRAFICO"):
                clave = "GRAFICO"
                principal = nombre
            elif "LUNES" in clave:
                clave = "LV"
            elif clave.replace("-", "").replace(" ", "") == "SDF":
                clave = "SDF"
            hojas[clave] = filas

        if "GRAFICO" not in hojas:
            print("  ! %s: no se encontro la hoja del grafico; se omite." % ruta.name)
            continue

        mes = mes_del_grafico(hojas, ruta.stem)
        if not mes:
            print("  ! %s: no se pudo deducir el mes; se omite." % ruta.name)
            continue

        graficos.append({"archivo": ruta.name, "mes": mes, "hoja": principal, "hojas": hojas})
        print("  + %s  ->  %s  (%s)" % (
            ruta.name, mes, ", ".join("%s:%d filas" % (k, len(v)) for k, v in hojas.items())))

    # Un grafico por mes: gana el ultimo modificado.
    unicos = {}
    for g in sorted(graficos, key=lambda g: (ORIGEN / g["archivo"]).stat().st_mtime):
        unicos[g["mes"]] = g
    graficos = sorted(unicos.values(), key=lambda g: g["mes"])

    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    DESTINO.write_text(
        json.dumps({"version": 1,
                    "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "graficos": graficos},
                   ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print("\nListo: %s  (%d graficos, %.0f KB)" % (
        DESTINO.relative_to(BASE), len(graficos), DESTINO.stat().st_size / 1024))
    print("Ahora sube al repositorio: grafico/grafico.json")


if __name__ == "__main__":
    main()
