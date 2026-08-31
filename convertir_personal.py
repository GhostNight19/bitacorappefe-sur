# -*- coding: utf-8 -*-
"""
Conversor de la dotacion de TRACCION (Excel) -> personal/personal.json

Uso:
    python convertir_personal.py

Deja el Excel de la dotacion dentro de "personal_excel" y ejecuta el script.

Se queda SOLO con el sector CONCEPCION y solo con quienes conducen: maquinistas
y ayudantes. Temuco y Puerto Montt quedan fuera, igual que los cargos que no
suben a la maquina (inspectores, jefaturas, supervisores).

Lo unico que la aplicacion necesita de aqui es saber si alguien es maquinista o
ayudante, para no preguntarle por horas de manejo a quien no maneja.
"""

import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    raise SystemExit("Falta openpyxl. Instalalo con:  pip install openpyxl")

BASE = Path(__file__).resolve().parent
ORIGEN = BASE / "personal_excel"
DESTINO = BASE / "personal" / "personal.json"

SECTOR = "CONCEPCION"

# Cargos que conducen. El resto de la dotacion no entra al archivo.
CARGOS = {
    "MAQUINISTA (P)": "MAQUINISTA",
    "MAQUINISTA INSTRUCTOR": "MAQUINISTA",
    "AYUDANTE DE MAQUINISTA": "AYUDANTE",
    "AYUDANTE DE MAQUINISTA/PATRULLERA": "AYUDANTE",
}


def sin_acentos(texto):
    return "".join(
        c for c in unicodedata.normalize("NFD", str(texto))
        if unicodedata.category(c) != "Mn"
    )


def norma(texto):
    return sin_acentos(texto).upper().strip()


def clave(nombre):
    """'J.MELO', 'J. MELO' y 'J. Melo' comparten la misma clave: JMELO."""
    return re.sub(r"[^A-Z0-9Ñ]", "", norma(nombre))


def columnas(cabecera):
    m = {}
    for i, c in enumerate(cabecera):
        t = norma(c)
        if t == "ABREVIACION":
            m["abreviacion"] = i
        elif t == "N-AA":
            m["nombre"] = i
        elif t == "SECTOR":
            m["sector"] = i
        elif t == "CARGO":
            m["cargo"] = i
        elif t == "ESTADO":
            m["estado"] = i
    return m


def main():
    if not ORIGEN.exists():
        ORIGEN.mkdir(parents=True)
        print('Se creo la carpeta "personal_excel". Deja ahi el Excel y vuelve a ejecutar.')
        return

    archivos = sorted(
        p for p in ORIGEN.iterdir()
        if p.suffix.lower() in (".xlsx", ".xlsm") and not p.name.startswith("~$")
    )
    if not archivos:
        print('No hay Excel en "personal_excel". Deja ahi la dotacion y vuelve a ejecutar.')
        return

    ruta = archivos[-1]
    ws = openpyxl.load_workbook(ruta, data_only=True)[
        openpyxl.load_workbook(ruta, data_only=True).sheetnames[0]]

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        print("El Excel esta vacio.")
        return

    cols = columnas(filas[0])
    faltan = [k for k in ("abreviacion", "sector", "cargo") if k not in cols]
    if faltan:
        print("Faltan columnas en el Excel: %s" % ", ".join(faltan))
        return

    gente = []
    descartes = {"otro sector": 0, "no conduce": 0, "sin abreviacion": 0}

    for fila in filas[1:]:
        val = lambda k: str(fila[cols[k]] or "").strip() if k in cols else ""
        abrev = val("abreviacion")
        if not abrev:
            descartes["sin abreviacion"] += 1
            continue
        if norma(val("sector")) != SECTOR:
            descartes["otro sector"] += 1
            continue
        rol = CARGOS.get(norma(val("cargo")))
        if not rol:
            descartes["no conduce"] += 1
            continue

        gente.append({
            "clave": clave(abrev),
            "abreviacion": abrev,
            "nombre": val("nombre") or abrev,
            "cargo": val("cargo"),
            "rol": rol,
            "estado": val("estado"),
        })

    gente.sort(key=lambda g: g["clave"])

    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    DESTINO.write_text(
        json.dumps({"version": 1,
                    "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "sector": SECTOR,
                    "archivo": ruta.name,
                    "personas": gente},
                   ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    maq = sum(1 for g in gente if g["rol"] == "MAQUINISTA")
    print("  + %s" % ruta.name)
    print("    %s: %d maquinistas + %d ayudantes = %d" % (SECTOR, maq, len(gente) - maq, len(gente)))
    print("    fuera: %d de otro sector, %d que no conducen, %d sin abreviacion" % (
        descartes["otro sector"], descartes["no conduce"], descartes["sin abreviacion"]))
    print("\nListo: %s  (%.0f KB)" % (DESTINO.relative_to(BASE), DESTINO.stat().st_size / 1024))
    print("Ahora sube al repositorio: personal/personal.json")


if __name__ == "__main__":
    main()
