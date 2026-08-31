# -*- coding: utf-8 -*-
"""
Conversor del BOLETIN DE VIA C (Excel) -> prevenciones/boletin.json

Uso:
    python convertir_boletin.py

Deja el Excel del boletin del dia dentro de la carpeta "boletines_excel" y
ejecuta el script. Genera "prevenciones/boletin.json", que es lo que lee la
aplicacion al abrirse (y lo que hay que subir al repositorio).

Igual que con las pautas, el script NO interpreta el boletin: vuelca la hoja
tal cual (filas de texto) y toda la lectura -tramos, PK, horarios, vias- la
hace el propio index.html.

Solo se conserva la parte util: desde "NOTIFICACION DE FAENAS EN EL INICIO DEL
RECORRIDO" hasta antes de "PROGRAMACION DE CORTADAS". Las cortadas son trabajos
con via ocupada fuera de servicio, no prevenciones de circulacion.
"""

import json
import re
import unicodedata
from datetime import date, datetime, time, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    raise SystemExit("Falta openpyxl. Instalalo con:  pip install openpyxl")

BASE = Path(__file__).resolve().parent
ORIGEN = BASE / "boletines_excel"
DESTINO = BASE / "prevenciones" / "boletin.json"

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}

CORTE = "PROGRAMACION DE CORTADAS"


def sin_acentos(texto):
    return "".join(
        c for c in unicodedata.normalize("NFD", str(texto))
        if unicodedata.category(c) != "Mn"
    )


def celda_a_texto(valor):
    """Texto plano, conservando horas (HH:MM) y fechas (AAAA-MM-DD)."""
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "SI" if valor else "NO"
    if isinstance(valor, datetime):
        if valor.hour or valor.minute or valor.second:
            if valor.year > 1900:
                return valor.strftime("%Y-%m-%d %H:%M")
            return valor.strftime("%H:%M")
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, time):
        return valor.strftime("%H:%M")
    if isinstance(valor, timedelta):
        total = int(valor.total_seconds() // 60)
        return "%d:%02d" % (total // 60, total % 60)
    if isinstance(valor, date):
        return valor.isoformat()
    if isinstance(valor, float):
        # Los PK vienen como 3.3, 27.00095918367347, 28.002...
        return ("%.5f" % valor).rstrip("0").rstrip(".") if not valor.is_integer() else str(int(valor))
    return str(valor).replace("\n", " ").strip()


def fecha_de_las_filas(filas, nombre):
    """Del encabezado ('Emitido 2026-08-27') o del nombre del archivo."""
    for fila in filas[:12]:
        for celda in fila:
            m = re.search(r"(20\d{2})-(\d{2})-(\d{2})", str(celda))
            if m:
                return m.group(0)
    m = re.search(r"(\d{2})-(\d{2})-(\d{2})\b", sin_acentos(nombre))
    if m:  # 27-08-26 -> 2026-08-27
        return "20%s-%s-%s" % (m.group(3), m.group(2), m.group(1))
    return ""


def folio_de_las_filas(filas, nombre):
    for i, fila in enumerate(filas[:12]):
        for j, celda in enumerate(fila):
            if sin_acentos(celda).strip().upper() == "FOLIO":
                for k in range(i + 1, min(i + 3, len(filas))):
                    v = str(filas[k][j] if j < len(filas[k]) else "").strip()
                    if v:
                        return v
    m = re.match(r"\s*(\d+[-A-Za-z]*)", nombre)
    return m.group(1) if m else ""


def orden_folio(folio):
    """243 < 243-A < 243-B. Sirve para saber cual es la ultima revision."""
    m = re.match(r"\s*(\d+)\s*-?\s*([A-Za-z]?)", str(folio))
    if not m:
        return (-1, "")
    return (int(m.group(1)), m.group(2).upper())


def leer_filas(ws):
    filas = []
    for fila in ws.iter_rows(values_only=True):
        celdas = [celda_a_texto(v) for v in fila]
        while celdas and celdas[-1] == "":
            celdas.pop()
        filas.append(celdas)

    # Se corta en "PROGRAMACION DE CORTADAS": lo de abajo no son prevenciones.
    for i, fila in enumerate(filas):
        if any(CORTE in sin_acentos(c).upper() for c in fila):
            filas = filas[:i]
            break

    while filas and not any(filas[-1]):
        filas.pop()
    return filas


def leer_hoja(ruta):
    """Un mismo Excel puede traer el boletin y sus reediciones en pestanas
    distintas (243, 243-A, 243-B). Manda siempre la ultima revision."""
    wb = openpyxl.load_workbook(ruta, data_only=True)

    candidatas = []
    for nombre in wb.sheetnames:
        filas = leer_filas(wb[nombre])
        folio = folio_de_las_filas(filas, "")
        if folio:
            candidatas.append((orden_folio(folio), nombre, folio, filas))

    if not candidatas:
        return leer_filas(wb[wb.sheetnames[0]]), ""

    candidatas.sort(key=lambda c: c[0])
    _, nombre, folio, filas = candidatas[-1]
    if len(candidatas) > 1:
        print("    (%d versiones en el archivo: %s -> se usa el folio %s)" % (
            len(candidatas), ", ".join(c[2] for c in candidatas), folio))
    return filas, folio


def main():
    if not ORIGEN.exists():
        ORIGEN.mkdir(parents=True)
        print('Se creo la carpeta "boletines_excel". Deja ahi el boletin y vuelve a ejecutar.')
        return

    archivos = sorted(
        p for p in ORIGEN.iterdir()
        if p.suffix.lower() in (".xlsx", ".xlsm") and not p.name.startswith("~$")
    )
    if not archivos:
        print('No hay Excel en "boletines_excel". Deja ahi el boletin y vuelve a ejecutar.')
        return

    hojas = []
    for ruta in archivos:
        filas, folio = leer_hoja(ruta)
        fecha = fecha_de_las_filas(filas, ruta.stem)
        if not fecha:
            print("  ! %s: no se pudo deducir la fecha; se omite." % ruta.name)
            continue
        hojas.append({
            "archivo": ruta.name,
            "fecha": fecha,
            "folio": folio or folio_de_las_filas(filas, ruta.stem),
            "filas": filas,
        })
        print("  + %s  ->  %s  folio %s  (%d filas utiles)" % (
            ruta.name, fecha, hojas[-1]["folio"], len(filas)))

    # Un boletin por fecha: gana la ultima revision del folio y, en empate,
    # el archivo copiado mas tarde.
    unicas = {}
    for hoja in sorted(hojas, key=lambda h: (orden_folio(h["folio"]),
                                             (ORIGEN / h["archivo"]).stat().st_mtime)):
        unicas[hoja["fecha"]] = hoja
    hojas = sorted(unicas.values(), key=lambda h: h["fecha"])

    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    DESTINO.write_text(
        json.dumps({"version": 1,
                    "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "boletines": hojas},
                   ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print("\nListo: %s  (%d boletines, %.0f KB)" % (
        DESTINO.relative_to(BASE), len(hojas), DESTINO.stat().st_size / 1024))
    print("Ahora sube al repositorio: prevenciones/boletin.json")


if __name__ == "__main__":
    main()
