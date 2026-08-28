# -*- coding: utf-8 -*-
"""
Conversor de pautas diarias (Excel) -> pautas/pautas.json

Uso:
    python convertir_pautas.py

Deja los Excel de las pautas dentro de la carpeta "pautas_excel" y ejecuta el
script. Genera "pautas/pautas.json", que es el archivo que lee la Bitacora al
abrirse (y el que hay que subir al repositorio junto con el index.html).

El script NO interpreta la pauta: solo vuelca la hoja tal cual (una lista de
filas con el texto de cada celda). Toda la lectura -turnos, tripulaciones,
descansos- la hace el propio index.html, para que el resultado sea identico
tanto si la pauta viene del repositorio como si se sube el Excel a mano desde
el celular.
"""

import json
import re
import unicodedata
from datetime import date, datetime, time, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    raise SystemExit("Falta openpyxl. Instalalo con:  pip install openpyxl")

BASE = Path(__file__).resolve().parent
ORIGEN = BASE / "pautas_excel"
DESTINO = BASE / "pautas" / "pautas.json"

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}


def sin_acentos(texto):
    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


def celda_a_texto(valor):
    """Convierte cualquier celda a texto plano, respetando las horas."""
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "SI" if valor else "NO"
    if isinstance(valor, datetime):
        return valor.strftime("%H:%M") if (valor.hour or valor.minute) else valor.strftime("%Y-%m-%d")
    if isinstance(valor, time):
        return valor.strftime("%H:%M")
    if isinstance(valor, timedelta):
        total = int(valor.total_seconds() // 60)
        return "%d:%02d" % (total // 60, total % 60)
    if isinstance(valor, date):
        return valor.isoformat()
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).replace("\n", " ").strip()


def fecha_de_las_filas(filas):
    """Busca 'viernes, 28 de agosto de 2026' en las primeras filas."""
    for fila in filas[:4]:
        for celda in fila:
            texto = sin_acentos(str(celda)).lower()
            m = re.search(r"(\d{1,2})\s+de\s+([a-z]+)\s+de\s+(\d{4})", texto)
            if m and m.group(2) in MESES:
                return date(int(m.group(3)), MESES[m.group(2)], int(m.group(1))).isoformat()
    return ""


def fecha_del_nombre(nombre):
    """Respaldo: 'Lunes 31 de agosto.xlsx' -> usa el anio en curso."""
    texto = sin_acentos(nombre).lower()
    m = re.search(r"(\d{1,2})\s*(?:de\s*)?([a-z]+)", texto)
    if not m or m.group(2) not in MESES:
        return ""
    anio = re.search(r"(20\d{2})", texto)
    anio = int(anio.group(1)) if anio else date.today().year
    try:
        return date(anio, MESES[m.group(2)], int(m.group(1))).isoformat()
    except ValueError:
        return ""


def leer_hoja(ruta):
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = []
    for fila in ws.iter_rows(values_only=True):
        celdas = [celda_a_texto(v) for v in fila]
        while celdas and celdas[-1] == "":
            celdas.pop()
        filas.append(celdas)
    while filas and not any(filas[-1]):
        filas.pop()
    return filas


def main():
    if not ORIGEN.exists():
        ORIGEN.mkdir(parents=True)
        print('Se creo la carpeta "pautas_excel". Deja ahi los Excel y vuelve a ejecutar.')
        return

    archivos = sorted(
        p for p in ORIGEN.iterdir()
        if p.suffix.lower() in (".xlsx", ".xlsm") and not p.name.startswith("~$")
    )
    if not archivos:
        print('No hay Excel en "pautas_excel". Deja ahi las pautas y vuelve a ejecutar.')
        return

    hojas = []
    for ruta in archivos:
        filas = leer_hoja(ruta)
        fecha = fecha_de_las_filas(filas) or fecha_del_nombre(ruta.stem)
        if not fecha:
            print("  ! %s: no se pudo deducir la fecha; se omite." % ruta.name)
            continue
        hojas.append({"archivo": ruta.name, "fecha": fecha, "filas": filas})
        print("  + %s  ->  %s  (%d filas)" % (ruta.name, fecha, len(filas)))

    # Si hay dos archivos con la misma fecha, gana el ultimo modificado.
    unicas = {}
    for hoja in sorted(hojas, key=lambda h: (ORIGEN / h["archivo"]).stat().st_mtime):
        unicas[hoja["fecha"]] = hoja
    hojas = sorted(unicas.values(), key=lambda h: h["fecha"])

    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    salida = {
        "version": 1,
        "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "hojas": hojas,
    }
    DESTINO.write_text(
        json.dumps(salida, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print("\nListo: %s  (%d pautas, %.0f KB)" % (
        DESTINO.relative_to(BASE), len(hojas), DESTINO.stat().st_size / 1024))
    print("Ahora sube al repositorio: pautas/pautas.json")


if __name__ == "__main__":
    main()
